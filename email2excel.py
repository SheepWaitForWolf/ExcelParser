import re
import openpyxl
from openpyxl.workbook import Workbook

wb = openpyxl.load_workbook('MOCK_DATA.xlsx')

sheet = wb['data']

# Create arrays to store data from input file
Names = []
Postcodes = []
DoBs = []
Genders = []

# Construct the regex pattern for matching postcodes
# post_pattern = "GIR[ ]?0AA|((AB|AL|B|BA|BB|BD|BH|BL|BN|BR|BS|BT|CA|CB|CF|CH|CM|CO|CR|CT|CV|CW|DA|DD|DE|DG|DH|DL|DN|DT|DY|E|EC|EH|EN|EX|FK|FY|G|GL|GY|GU|HA|HD|HG|HP|HR|HS|HU|HX|IG|IM|IP|IV|JE|KA|KT|KW|KY|L|LA|LD|LE|LL|LN|LS|LU|M|ME|MK|ML|N|NE|NG|NN|NP|NR|NW|OL|OX|PA|PE|PH|PL|PO|PR|RG|RH|RM|S|SA|SE|SG|SK|SL|SM|SN|SO|SP|SR|SS|ST|SW|SY|TA|TD|TF|TN|TQ|TR|TS|TW|UB|W|WA|WC|WD|WF|WN|WR|WS|WV|YO|ZE)(\d[\dA-Z]?[ ]?\d[ABD-HJLN-UW-Z]{2}))|BFPO[ ]?\d{1,4}"

n = sheet.max_row

# Read from Excel file
for i in range(2, n): 
	Names.append(sheet.cell(row = i, column = 1).value)
	DoBs.append(sheet.cell(row = i, column = 4).value)
	Genders.append(sheet.cell(row = i, column = 5).value)	
	Postcodes.append(sheet.cell(row = i, column = 2).value)


wbout = openpyxl.load_workbook('Out.xlsx')
outsheet = wbout['Sheet1']


# Populate output file with common data

for i in range(3,n):
	outsheet[i] = outsheet[2]
wbout.save('Out.xlsx')

# Filter Addresses down to Postcodes
# for i in range(0,Postcodes.length()):
# 	re.search


# Write to Excel output file
for i in range(0, len(Names)):
		outsheet.cell(row=(2 + i), column=5).value = Names[i]
		outsheet.cell(row=(2 + i), column=6).value = Genders[i]
		outsheet.cell(row=(2 + i), column=8).value = DoBs[i]


wbout.save('Out.xlsx')
print('Finito!')