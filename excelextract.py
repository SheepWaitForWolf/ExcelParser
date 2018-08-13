import re
import openpyxl
from openpyxl.workbook import Workbook

# Load input and output files
wb = openpyxl.load_workbook('MOCK_DATA.xlsx')
wbout = openpyxl.load_workbook('Out.xlsx')

# Put worksheets in working memory
sheet = wb['data']
outsheet = wbout['Sheet1']

n = sheet.max_row 

# Create arrays to store data from input file
Names = []
Addresses = []
DoBs = []
Genders = []
Postcodes = []
ListOfLists = []

# Read from Excel file and populate lists with Data
for i in range(2, n+1): 
	Names.append(sheet.cell(row = i, column = 1).value)
	Addresses.append(sheet.cell(row = i, column = 2).value)	
	DoBs.append(sheet.cell(row = i, column = 4).value)
	Genders.append(sheet.cell(row = i, column = 5).value)

# Loop through Address list, split strings into lists 
for i in range(n-1):
	ListOfLists.append(Addresses[i].split())

# Loop through ListofLists and combine last two elements into Postcodes list	
for i in range(n-1):
	Postcodes.append(ListOfLists[i][-2] + " " + ListOfLists[i][-1])

# Write to Excel output file
for i in range(0, len(Names)):
		outsheet.cell(row=(2 + i), column=5).value = Names[i]
		outsheet.cell(row=(2 + i), column=6).value = Genders[i]
		outsheet.cell(row=(2 + i), column=7).value = Postcodes[i]
		outsheet.cell(row=(2 + i), column=8).value = DoBs[i]

wbout.save('Out.xlsx')
print('Finito!')